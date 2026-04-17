#!/usr/bin/env python3
"""
Keeper — Книга Учёта Доходов и Расходов
TUI-приложение для ведения бизнеса в стиле neovim
"""

from __future__ import annotations

import json
import os
from datetime import date, datetime
from pathlib import Path
from typing import cast

from textual import on
from textual.app import App, ComposeResult
from textual.binding import Binding
from textual.containers import Container, Horizontal, Vertical, Center
from textual.screen import ModalScreen
from textual.widget import Widget
from textual.widgets import (
    DataTable,
    Footer,
    Header,
    Input,
    Label,
    Button,
    Static,
    Switch,
)
from textual.widgets.data_table import RowKey

import openpyxl
from openpyxl.workbook import Workbook
from dateutil.parser import parse as parse_date


# Цветовая схема Dracula-inspired
COLORS = {
    "bg": "#1e1e2e",
    "surface": "#313244",
    "primary": "#89b4fa",
    "secondary": "#a6e3a1",  # доходы
    "tertiary": "#f38ba8",    # расходы
    "text": "#cdd6f4",
    "text_secondary": "#a6adc8",
    "border": "#45475a",
}


class Transaction:
    """Модель транзакции"""
    def __init__(self, date: str, name: str, amount: float, type_: str):
        self.date = date
        self.name = name
        self.amount = amount
        self.type_ = type_  # "IN" или "OUT"


class AddTransactionModal(ModalScreen[dict | None]):
    """Модальное окно добавления/редактирования транзакции"""

    BINDINGS = [
        Binding("escape", "cancel", "Отмена"),
        Binding("enter", "submit", "Сохранить"),
    ]

    def __init__(self, transaction: Transaction | None = None) -> None:
        super().__init__()
        self.transaction = transaction

    def compose(self) -> ComposeResult:
        title = "Редактирование" if self.transaction else "Новая транзакция"

        with Container(id="modal-container"):
            with Vertical(id="modal-content"):
                yield Static(title, id="modal-title")

                yield Label("Дата (ГГГГ-ММ-ДД):", id="date-label")
                default_date = self.transaction.date if self.transaction else date.today().isoformat()
                yield Input(value=default_date, id="date-input", placeholder="2024-01-15")

                yield Label("Название:", id="name-label")
                yield Input(
                    value=self.transaction.name if self.transaction else "",
                    id="name-input",
                    placeholder="Описание операции",
                    max_length=100
                )

                yield Label("Сумма:", id="amount-label")
                yield Input(
                    value=str(self.transaction.amount) if self.transaction else "",
                    id="amount-input",
                    placeholder="0.00",
                    type="number"
                )

                yield Label("Тип:", id="type-label")
                with Horizontal(id="type-toggle"):
                    yield Label("РАСХОД", id="income-label")
                    is_income = self.transaction.type_ == "IN" if self.transaction else True
                    yield Switch(value=is_income, id="type-switch")
                    yield Label("ДОХОД", id="expense-label")

                with Horizontal(id="modal-buttons"):
                    yield Button("Отмена", variant="default", id="cancel-btn")
                    yield Button("Сохранить", variant="primary", id="save-btn")

    def action_cancel(self) -> None:
        self.dismiss(None)

    def action_submit(self) -> None:
        self._save_transaction()

    @on(Button.Pressed, "#cancel-btn")
    def on_cancel(self) -> None:
        self.dismiss(None)

    @on(Button.Pressed, "#save-btn")
    def on_save(self) -> None:
        self._save_transaction()

    def _save_transaction(self) -> None:
        date_input = self.query_one("#date-input", Input)
        name_input = self.query_one("#name-input", Input)
        amount_input = self.query_one("#amount-input", Input)
        type_switch = self.query_one("#type-switch", Switch)

        # Валидация
        try:
            parsed_date = parse_date(date_input.value).date()
            date_str = parsed_date.isoformat()
        except (ValueError, TypeError):
            self.notify("Некорректная дата. Используйте формат ГГГГ-ММ-ДД", severity="error")
            return

        name = name_input.value.strip()
        if not name:
            self.notify("Введите название транзакции", severity="error")
            return

        try:
            amount = float(amount_input.value)
            if amount <= 0:
                raise ValueError
        except (ValueError, TypeError):
            self.notify("Введите корректную сумму (> 0)", severity="error")
            return

        type_ = "ДОХОД" if type_switch.value else "РАСХОД"

        self.dismiss({
            "date": date_str,
            "name": name,
            "amount": amount,
            "type": type_
        })


class SplitModal(ModalScreen[str | None]):
    """Модальное окно создания нового файла (split)"""

    BINDINGS = [
        Binding("escape", "cancel", "Отмена"),
        Binding("enter", "submit", "Создать"),
    ]

    def compose(self) -> ComposeResult:
        with Container(id="modal-container"):
            with Vertical(id="modal-content"):
                yield Static("Новый период", id="modal-title")

                yield Label("Имя файла:", id="filename-label")
                current_date = datetime.now().strftime("%Y_%m")
                yield Input(
                    value=f"keeper_{current_date}.xlsx",
                    id="filename-input",
                    placeholder="keeper_2024_01.xlsx"
                )

                yield Label(
                    "Текущий файл останется без изменений.\n"
                    "Будет создан новый Excel-файл.",
                    id="split-hint"
                )

                with Horizontal(id="modal-buttons"):
                    yield Button("Отмена", variant="default", id="cancel-btn")
                    yield Button("Создать", variant="primary", id="create-btn")

    def action_cancel(self) -> None:
        self.dismiss(None)

    def action_submit(self) -> None:
        self._create_split()

    @on(Button.Pressed, "#cancel-btn")
    def on_cancel(self) -> None:
        self.dismiss(None)

    @on(Button.Pressed, "#create-btn")
    def on_create(self) -> None:
        self._create_split()

    def _create_split(self) -> None:
        filename_input = self.query_one("#filename-input", Input)
        filename = filename_input.value.strip()

        if not filename:
            self.notify("Введите имя файла", severity="error")
            return

        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        self.dismiss(filename)


class KeeperApp(App):
    """Основное приложение Keeper"""

    TITLE = "KEEPER"
    import random
    SUB_TITLE = random.choice(list(open('splash', 'r', encoding='utf-8'))).strip()

    CSS = """
    Screen {
        background: $surface;
    }

    #main-container {
        height: 100%;
        padding: 1 2;
    }

    #file-info {
        height: 3;
        padding: 1 2;
        background: $primary;
        color: #1e1e2e;
        text-style: bold;
        margin-bottom: 1;
    }

    #transaction-table {
        height: 1fr;
        background: #1e1e2e;
        border: solid $border;
    }

    DataTable {
        background: #1e1e2e;
        color: $text;
    }

    DataTable > .datatable--header {
        background: $primary;
        color: #1e1e2e;
        text-style: bold;
    }

    DataTable > .datatable--cursor {
        background: $primary 50%;
        color: $text;
    }

    DataTable > .datatable--hover {
        background: $surface;
    }

    #totals-bar {
        height: 3;
        margin-top: 1;
        padding: 1 2;
        background: #1e1e2e;
        border: solid $border;
    }

    #totals-bar Label {
        color: $text;
        padding: 0 1;
    }

    #income-total {
        color: $secondary;
        text-style: bold;
    }

    #expense-total {
        color: #f38ba8;
        text-style: bold;
    }

    #balance-total {
        color: $primary;
        text-style: bold;
    }

    /* Modal styles */
    #modal-container {
        align: center middle;
        width: 100%;
        height: 100%;
        background: #1e1e2e 80%;
    }

    #modal-content {
        width: 60;
        height: auto;
        background: $surface;
        border: solid $primary;
        padding: 2 3;
    }

    #modal-title {
        text-style: bold;
        color: $primary;
        text-align: center;
        padding-bottom: 1;
    }

    #modal-content Label {
        color: #a6adc8;
        margin-top: 1;
    }

    #modal-content Input {
        width: 100%;
        margin-bottom: 1;
    }

    #type-toggle {
        width: 100%;
        height: 3;
        align: center middle;
        margin: 1 0;
    }

    #type-toggle Label {
        width: 1fr;
        text-align: center;
        margin: 0 1;
    }

    #income-label {
        color: $secondary;
    }

    #expense-label {
        color: #f38ba8;
    }

    #modal-buttons {
        width: 100%;
        height: 3;
        align: center middle;
        margin-top: 2;
    }

    #modal-buttons Button {
        width: 1fr;
        margin: 0 1;
    }

    #split-hint {
        color: #a6adc8;
        text-align: center;
        padding: 1 0;
        margin: 1 0;
    }

    #custom-footer {
        height: 3;
        padding: 1 2;
        background: $primary;
        color: #1e1e2e;
        text-align: center;
        text-style: bold;
    }
    """

    BINDINGS = [
        Binding("q", "quit", "Выход"),
        Binding("й", "quit", "Выход"),
        Binding("a", "add_transaction", "Добавить"),
        Binding("ф", "add_transaction", "Добавить"),
        Binding("e", "edit_transaction", "Редактировать"),
        Binding("у", "edit_transaction", "Редактировать"),
        Binding("d", "delete_transaction", "Удалить"),
        Binding("в", "delete_transaction", "Удалить"),
        Binding("s", "split_period", "Разделить"),
        Binding("ы", "split_period", "Разделить"),
        Binding("o", "open_file", "Открыть"),
        Binding("щ", "open_file", "Открыть"),
        Binding("n", "new_file", "Новый"),
        Binding("т", "new_file", "Новый"),
       # Binding("j", "cursor_down", "Вниз"),
       # Binding("k", "cursor_up", "Вверх"),
    ]

    def __init__(self) -> None:
        super().__init__()
        self.current_file: Path | None = None
        self.transactions: list[Transaction] = []
        self.sheet_name = "Транзакции"
        self.config_file = Path.home() / ".keeper_last_file.json"

    def compose(self) -> ComposeResult:
        yield Header()

        with Container(id="main-container"):
            yield Static("Файл: не выбран", id="file-info")
            yield DataTable(id="transaction-table")

            with Horizontal(id="totals-bar"):
                yield Label(f"Доходы: ", id="income-label")
                yield Label("0.00", id="income-total")
                yield Label(" | Расходы: ", id="expense-label")
                yield Label("0.00", id="expense-total")
                yield Label(" | Баланс: ", id="balance-label")
                yield Label("0.00", id="balance-total")

        yield Static("", id="custom-footer")

    def on_mount(self) -> None:
        """Инициализация при запуске"""
        table = self.query_one("#transaction-table", DataTable)
        table.add_columns("Дата", "Название", "Сумма", "Тип")
        table.zebra_stripes = True
        table.focus()

        # Сначала пробуем открыть последний файл
        last_file = self._get_last_file()
        if last_file and last_file.exists():
            self.load_file(last_file)
            self.notify(f"Открыт последний файл: {last_file.name}")
        else:
            # Если последнего файла нет, пробуем дефолтный
            default_file = Path("keeper.xlsx")
            if default_file.exists():
                self.load_file(default_file)
            else:
                self.notify("Создайте новый файл (n) или откройте существующий (o)")

    def action_cursor_down(self) -> None:
        table = self.query_one("#transaction-table", DataTable)
        table.cursor_down()

    def action_cursor_up(self) -> None:
        table = self.query_one("#transaction-table", DataTable)
        table.cursor_up()

    def action_add_transaction(self) -> None:
        """Добавить новую транзакцию"""
        if not self.current_file:
            self.notify("Сначала откройте или создайте файл", severity="warning")
            return

        def handle_result(result: dict | None) -> None:
            if result:
                self._add_transaction_to_sheet(result)
                self.refresh_transactions()

        self.push_screen(AddTransactionModal(), handle_result)

    def action_edit_transaction(self) -> None:
        """Редактировать выбранную транзакцию"""
        if not self.current_file:
            self.notify("Сначала откройте или создайте файл", severity="warning")
            return

        table = self.query_one("#transaction-table", DataTable)
        if table.cursor_row is None:
            self.notify("Выберите транзакцию для редактирования", severity="warning")
            return

        # Получаем данные транзакции
        transaction = self.transactions[table.cursor_row]

        def handle_result(result: dict | None) -> None:
            if result:
                self._update_transaction_in_sheet(table.cursor_row, result)
                self.refresh_transactions()

        self.push_screen(AddTransactionModal(transaction), handle_result)

    def action_delete_transaction(self) -> None:
        """Удалить выбранную транзакцию"""
        if not self.current_file:
            self.notify("Сначала откройте или создайте файл", severity="warning")
            return

        table = self.query_one("#transaction-table", DataTable)
        if table.cursor_row is None:
            self.notify("Выберите транзакцию для удаления", severity="warning")
            return

        row_index = table.cursor_row
        transaction = self.transactions[row_index]

        def confirm_delete() -> None:
            self._delete_transaction_from_sheet(row_index)
            self.refresh_transactions()
            self.notify("Транзакция удалена")

       # self.notify(
       #     f"Удалить \"{transaction.name}\"?",
       #     severity="information",
       #     timeout=None,
       #     buttons=[("Да", confirm_delete), ("Нет", lambda: None)]
       # )
        confirm_delete()
    def action_split_period(self) -> None:
        """Создать новый файл для нового периода"""
        def handle_result(filename: str | None) -> None:
            if filename:
                new_path = Path(filename)
                if new_path.exists():
                    self.notify(f"Файл {filename} уже существует", severity="error")
                    return

                # Создаём новый пустой файл
                self._create_excel_file(new_path)
                self.load_file(new_path)
                self.notify(f"Создан новый файл: {filename}")

        self.push_screen(SplitModal(), handle_result)

    def action_open_file(self) -> None:
        """Открыть существующий файл"""
        self._open_file_dialog()

    def _open_file_dialog(self) -> None:
        """Диалог открытия файла"""
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)

        file_path = filedialog.askopenfilename(
            title="Открыть файл Keeper",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if file_path:
            self.load_file(Path(file_path))
            self.notify(f"Открыт файл: {Path(file_path).name}")

    def action_new_file(self) -> None:
        """Создать новый файл"""
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)

        file_path = filedialog.asksaveasfilename(
            title="Создать новый файл",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )

        if file_path:
            path = Path(file_path)
            self._create_excel_file(path)
            self.load_file(path)
            self.notify(f"Создан новый файл: {path.name}")

    def load_file(self, path: Path) -> None:
        """Загрузить данные из Excel файла"""
        try:
            wb = openpyxl.load_workbook(path)

            if self.sheet_name not in wb.sheetnames:
                # Создаём лист если не существует
                ws = wb.create_sheet(self.sheet_name)
                self._write_headers(ws)
            else:
                ws = wb[self.sheet_name]

            self.current_file = path
            self._save_last_file(path)
            self._update_file_info()
            self.refresh_transactions()

        except Exception as e:
            self.notify(f"Ошибка загрузки файла: {e}", severity="error")

    def _save_last_file(self, path: Path) -> None:
        """Сохранить путь к последнему файлу"""
        try:
            with open(self.config_file, "w", encoding="utf-8") as f:
                json.dump({"last_file": str(path)}, f)
        except Exception:
            pass  # Тихо игнорируем ошибки сохранения конфига

    def _get_last_file(self) -> Path | None:
        """Получить путь к последнему файлу"""
        try:
            if self.config_file.exists():
                with open(self.config_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    return Path(data.get("last_file", ""))
        except Exception:
            pass
        return None

    def refresh_transactions(self) -> None:
        """Обновить таблицу транзакций"""
        if not self.current_file:
            return

        try:
            wb = openpyxl.load_workbook(self.current_file)
            ws = wb[self.sheet_name]

            self.transactions = []
            table = self.query_one("#transaction-table", DataTable)
            table.clear()

            # Читаем данные (пропускаем заголовок)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                if row[0] is None:  # Пустая строка
                    continue

                date_val = row[0]
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%Y-%m-%d")
                elif isinstance(date_val, date):
                    date_str = date_val.isoformat()
                else:
                    date_str = str(date_val)

                name = str(row[1]) if row[1] else ""
                amount = float(row[2]) if row[2] else 0.0
                type_ = str(row[3]) if row[3] else "ДОХОД"

                transaction = Transaction(date_str, name, amount, type_)
                self.transactions.append(transaction)

            # Сортируем по дате (новые сверху) ПЕРЕД добавлением в таблицу
            self.transactions.sort(key=lambda t: t.date, reverse=True)

            # Добавляем отсортированные данные в таблицу
            for idx, transaction in enumerate(self.transactions):
                type_style = "[green]ДОХОД[/green]" if transaction.type_ == "ДОХОД" else "[red]РАСХОД[/red]"
                table.add_row(
                    transaction.date,
                    transaction.name,
                    f"{transaction.amount:,.2f}",
                    type_style,
                    key=f"row_{idx}"
                )

            self._update_totals()

        except Exception as e:
            self.notify(f"Ошибка чтения данных: {e}", severity="error")

    def _update_totals(self) -> None:
        """Обновить итоговые суммы"""
        income = sum(t.amount for t in self.transactions if t.type_ == "ДОХОД")
        expense = sum(t.amount for t in self.transactions if t.type_ == "РАСХОД")
        balance = income - expense

        self.query_one("#income-total", Label).update(f"{income:,.2f}")
        self.query_one("#expense-total", Label).update(f"{expense:,.2f}")
        self.query_one("#balance-total", Label).update(f"{balance:,.2f}")
        self._update_footer(balance)

    def _update_footer(self, balance: float) -> None:
        """Обновить кастомный footer"""
        footer = self.query_one("#custom-footer", Static)
        balance_str = f"{balance:,.2f}"
        balance_color = "[green]" if balance >= 0 else "[red]"
        footer.update(
            f"n — Новый  |  o — Открыть  |  a — Добавить  |  e — Edit  |  d — Удалить  |  "
            f"s — Split  |  q — Выход     {balance_color}● Баланс: {balance_str}[/]"
        )

    def _update_file_info(self) -> None:
        """Обновить информацию о текущем файле"""
        file_info = self.query_one("#file-info", Static)
        if self.current_file:
            file_info.update(f"📁 {self.current_file.name}")
        else:
            file_info.update("Файл: не выбран")

    def _add_transaction_to_sheet(self, data: dict) -> None:
        """Добавить транзакцию в Excel"""
        if not self.current_file:
            return

        try:
            wb = openpyxl.load_workbook(self.current_file)
            ws = wb[self.sheet_name]

            # Находим первую пустую строку
            next_row = ws.max_row + 1

            ws.cell(row=next_row, column=1, value=data["date"])
            ws.cell(row=next_row, column=2, value=data["name"])
            ws.cell(row=next_row, column=3, value=data["amount"])
            ws.cell(row=next_row, column=4, value=data["type"])

            wb.save(self.current_file)
            self.notify("Транзакция добавлена")

        except Exception as e:
            self.notify(f"Ошибка сохранения: {e}", severity="error")

    def _update_transaction_in_sheet(self, row_index: int, data: dict) -> None:
        """Обновить транзакцию в Excel"""
        if not self.current_file:
            return

        try:
            wb = openpyxl.load_workbook(self.current_file)
            ws = wb[self.sheet_name]

            # +2 т.к. индекс 0-based, а строка 1-based + заголовок
            excel_row = row_index + 2

            ws.cell(row=excel_row, column=1, value=data["date"])
            ws.cell(row=excel_row, column=2, value=data["name"])
            ws.cell(row=excel_row, column=3, value=data["amount"])
            ws.cell(row=excel_row, column=4, value=data["type"])

            wb.save(self.current_file)
            self.notify("Транзакция обновлена")

        except Exception as e:
            self.notify(f"Ошибка сохранения: {e}", severity="error")

    def _delete_transaction_from_sheet(self, row_index: int) -> None:
        """Удалить транзакцию из Excel"""
        if not self.current_file:
            return

        try:
            wb = openpyxl.load_workbook(self.current_file)
            ws = wb[self.sheet_name]

            # +2 т.к. индекс 0-based, а строка 1-based + заголовок
            excel_row = row_index + 2
            ws.delete_rows(excel_row)

            wb.save(self.current_file)

        except Exception as e:
            self.notify(f"Ошибка удаления: {e}", severity="error")

    def _create_excel_file(self, path: Path) -> None:
        """Создать новый Excel файл с заголовками"""
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        self._write_headers(ws)
        wb.save(path)

    def _write_headers(self, ws) -> None:
        """Записать заголовки в лист"""
        headers = ["Дата", "Название", "Сумма", "Тип"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # Стилизация заголовков
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="89B4FA", end_color="89B4FA", fill_type="solid")
            cell.font = openpyxl.styles.Font(color="1E1E2E", bold=True)


def main():
    app = KeeperApp()
    app.run()


if __name__ == "__main__":
    main()
